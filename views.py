from django.shortcuts import render
from django.shortcuts import redirect
from django.contrib import messages
from .models import AdminUser
from django.contrib.auth.hashers import make_password, check_password
from django.shortcuts import  get_object_or_404
from .models import Book
from django.contrib.auth import authenticate, login
def admin_login(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']

        try:
            admin = AdminUser.objects.get(username=username)
            if check_password(password, admin.password):
                request.session['admin'] = admin.username
                return redirect('dashboard')
            else:
                messages.error(request, "Invalid password")
        except AdminUser.DoesNotExist:
            messages.error(request, "Admin not found")
            
    return render(request, "login.html")
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.hashers import check_password
 # import your custom model

def dashboard(request):
    if 'admin' not in request.session:
        return redirect('admin_login')
    return render(request, "dashboard.html")
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from .models import AdminUser










def create_admin(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        confirm_password = request.POST['confirm_password']

        if password != confirm_password:
            messages.error(request, "Passwords do not match")
        elif AdminUser.objects.filter(username=username).exists():
            messages.error(request, "Username already exists")
        else:
            admin = AdminUser(username=username, password=make_password(password))
            admin.save()
            messages.success(request, "Admin created successfully")
            return redirect('admin_login')

    return render(request, "create_admin.html")
def index(request):
    return render(request,"index.html")


# Page to add a new book
# Page to add a new book
def add_book(request):
    if request.method == "POST":
        scanner = request.POST['scanner']
        title = request.POST['title']
        author = request.POST['author']
        year = int(request.POST['year'])
        publisher = request.POST['publisher']
        abstract = request.POST.get('abstract','') 
        department = request.POST.get('department')
        # If scanner already exists, just update the details
        book, created = Book.objects.get_or_create(
            scanner=scanner,
            defaults={
                'title': title,
                'author': author,
                'year': year,
                'publisher': publisher,
                'abstract': abstract,
                'available': True   # Default new book = Available
            }
        )

        if not created:
            # Already exists → update details only
            book.title = title
            book.author = author
            book.year = year
            book.publisher = publisher
            book.abstract = abstract
            book.department= department
            book.available = True  # Book re-added = available
            book.save()

        return redirect('available')  # redirect to function name

    return render(request, 'add_book.html')

from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Student, Book   # <-- Import both models

# Student Login
def student_log(request):
    if request.method == "POST":
        student_id = request.POST.get("student_id")
        password = request.POST.get("password")

        # Check if student exists and password is same as ID
        if student_id == password:
            if Student.objects.filter(student_id=student_id).exists():
                request.session["student_id"] = student_id
                return redirect("available")   # success → show books
            else:
                messages.error(request, "❌ College ID not found in database.")
        else:
            messages.error(request, "❌ Username and Password must be the same College ID.")

    return render(request, "student_logi.html")


# Show available books
def available_books(request):
    books = Book.objects.all()

    for b in books:
        # ✅ check Borrow table by matching book_id
        b.is_currently_available = not Borrow.objects.filter(
            book_id=b.scanner,
            returned=False    # still not returned
        ).exists()

    return render(request, 'available.html', {'books': books})


# Borrow a book

from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import Student

def scan_page(request):
    return render(request, "scanner.html")

def save_student(request):
    if request.method == "POST":
        student_id = request.POST.get("student_id")
        student_name = request.POST.get("student_name")

        Student.objects.create(student_id=student_id, student_name=student_name)
        return HttpResponse("✅ Student Saved Successfully")
    return redirect("scan_page")
def student_detail(request, student_id):
    student = Student.objects.get(id=student_id)
    return render(request, "student_detail.html", {"student": student})

from django.shortcuts import render, redirect

def student_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        # check credentials manually
        if username == "admin" and password == "ciet":
            return redirect("scan")   # go to scan page
        else:
            return render(request, "student_login.html", {"error": "Invalid Username or Password"})

    return render(request, "student_login.html")


def books_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        # check credentials manually
        if username == "admin" and password == "ciet":
            return redirect("add_book")   # go to scan page
        else:
            return render(request, "book_login.html", {"error": "Invalid Username or Password"})

    return render(request, "book_login.html")

def delete_login(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        # check credentials manually
        if username == "admin" and password == "ciet":
            return redirect("delete_book")   # go to scan page
        else:
            return render(request, "delete_login.html", {"error": "Invalid Username or Password"})
   
    return render(request, "delete_login.html")

def deletes_stud(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        # check credentials manually
        if username == "admin" and password == "ciet":
            return redirect("delete_student")   # go to scan page
        else:
            return render(request, "deletes_stud.html", {"error": "Invalid Username or Password"})

    return render(request, "deletes_stud.html")
from .models import Borrow, Student

# STEP 1: Scan Student
def borrow(request):
    if request.method == "POST":
        student_id = request.POST.get("student_id")
        student_name = request.POST.get("student_name")

        exists = Student.objects.filter(student_id=student_id, student_name=student_name).exists()

        if exists:
            request.session["student_id"] = student_id
            request.session["student_name"] = student_name
            return redirect("scan_book")
        else:
            return render(request, "borrow.html", {"error": "❌ Student not found in database"})

    return render(request, "borrow.html")

from django.utils import timezone

# STEP 2: Scan Book & Save Borrow
def scan_book(request):
    if request.method == "POST":
        book_id = request.POST.get("scanner")
        book_name = request.POST.get("title")

        student_id = request.session.get("student_id")
        student_name = request.session.get("student_name")

        # 🔹 Check if same student already borrowed the same book
        record = Borrow.objects.filter(
            student_id=student_id,
            book_id=book_id,
            returned=False
        ).first()

        if record:
            # ✅ Mark as returned instead of creating new row
            record.returned = True
            record.returned_at = timezone.now()
            record.save()
        else:
            # ✅ Create new borrow record
            Borrow.objects.create(
                student_id=student_id,
                student_name=student_name,
                book_id=book_id,
                book_name=book_name
            )

        return redirect("borrow_list")

    return render(request, "scan_book.html")
# STEP 3: Borrow List
def borrow_list(request):
    records = Borrow.objects.all().order_by('borrowed_at')
    

    # Dynamically compute borrow duration in days
    for b in records:
        if b.returned and b.returned_at:
            delta = b.returned_at - b.borrowed_at
            b.borrow_duration_days = delta.days  # integer days
        else:
            b.borrow_duration_days = None  # not returned yet

    return render(request, 'borrow_list.html', {'records': records})

    

# STEP 4: Mark as Returned
def return_book(request, borrow_id):
    borrow = Borrow.objects.get(id=borrow_id)
    borrow.returned = True   # ✅ FIXED
    borrow.save()
    return redirect("borrow_list")



def show(request):
    data=Book.objects.all()
    return render(request,'delete_book.html',{'data':data})

def delete(request,idn):
    obj=Book.objects.get(id=idn)
    obj.delete()
    return redirect('/show/')




def show1(request):
    data=Student.objects.all()
    return render(request,'delete_student.html',{'data':data})

def delete1(request,idn1):
    obj=Student.objects.get(id=idn1)
    obj.delete()
    return redirect('/show1/')




# yourapp/views.py
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from django.contrib.auth.decorators import login_required
from openpyxl import load_workbook
from io import BytesIO
import csv, io

from .forms import UploadFileForm
from .models import Student, Book

#@login_required
def bulk_upload_students(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            f = request.FILES['file']
            created, updated, skipped = 0, 0, 0
            try:
                if f.name.endswith('.xlsx'):
                    wb = load_workbook(filename=BytesIO(f.read()), data_only=True)
                    ws = wb.active
                    rows = ws.iter_rows(values_only=True)
                    headers = [str(h).strip() for h in next(rows)]
                    with transaction.atomic():
                        for row in rows:
                            data = dict(zip(headers, row))
                            sid = str(data.get('student_id'))
                            if not sid:
                                skipped += 1
                                continue
                            defaults = {'student_name': data.get('student_name') or ''}
                            obj, created_flag = Student.objects.update_or_create(
                                student_id=sid, defaults=defaults
                            )
                            if created_flag: created += 1
                            else: updated += 1

                elif f.name.endswith('.csv'):
                    text = f.read().decode('utf-8')
                    reader = csv.DictReader(io.StringIO(text))
                    with transaction.atomic():
                        for data in reader:
                            sid = str(data.get('student_id'))
                            if not sid:
                                skipped += 1
                                continue
                            defaults = {'student_name': data.get('student_name') or ''}
                            obj, created_flag = Student.objects.update_or_create(
                                student_id=sid, defaults=defaults
                            )
                            if created_flag: created += 1
                            else: updated += 1
                else:
                    messages.error(request, "Upload .xlsx or .csv only")
                    return redirect('bulk_upload_students')

                messages.success(request, f"Students Imported → created:{created}, updated:{updated}, skipped:{skipped}")
                return redirect('dashboard')# adjust to your actual page
            except Exception as e:
                messages.error(request, f"Error: {e}")
                return redirect('bulk_upload_students')
    else:
        form = UploadFileForm()
    return render(request, "bulk_upload_students.html", {"form": form})


# yourapp/views.py
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from django.contrib.auth.decorators import login_required
from openpyxl import load_workbook
from io import BytesIO
import csv, io

from .forms import UploadFileForm
from .models import Student, Book



#@login_required
def bulk_upload_books(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            f = request.FILES['file']
            created, updated, skipped = 0, 0, 0
            try:
                if f.name.endswith('.xlsx'):
                    wb = load_workbook(filename=BytesIO(f.read()), data_only=True)
                    ws = wb.active
                    rows = ws.iter_rows(values_only=True)
                    headers = [str(h).strip() for h in next(rows)]
                    with transaction.atomic():
                        for row in rows:
                            data = dict(zip(headers, row))
                            scanner = str(data.get('scanner') or '')  # Use scanner column
                            if not scanner:
                                skipped += 1
                                continue
                            defaults = {
                                'title': data.get('title') or '',
                                'author': data.get('author') or '',
                                'year': data.get('year') or None,
                                'edition': data.get('edition') or '',
                                'abstract': data.get('abstract') or '',
                            }
                            obj, created_flag = Book.objects.update_or_create(
                                scanner=scanner,  # 🔑 Corrected field
                                defaults=defaults
                            )
                            if created_flag:
                                created += 1
                            else:
                                updated += 1

                elif f.name.endswith('.csv'):
                    text = f.read().decode('utf-8')
                    reader = csv.DictReader(io.StringIO(text))
                    with transaction.atomic():
                        for data in reader:
                            scanner = str(data.get('scanner') or '')
                            if not scanner:
                                skipped += 1
                                continue
                            defaults = {
                                'title': data.get('title') or '',
                                'author': data.get('author') or '',
                                'year': data.get('year') or None,
                                'edition': data.get('edition') or '',
                            }
                            obj, created_flag = Book.objects.update_or_create(
                                scanner=scanner,  # 🔑 Corrected field
                                defaults=defaults
                            )
                            if created_flag:
                                created += 1
                            else:
                                updated += 1
                else:
                    messages.error(request, "Upload .xlsx or .csv only")
                    return redirect('bulk_upload_books')

                messages.success(request, f"Books Imported → created:{created}, updated:{updated}, skipped:{skipped}")
                return redirect('dashboard')
            except Exception as e:
                messages.error(request, f"Error: {e}")
                return redirect('bulk_upload_books')
    else:
        form = UploadFileForm()
    return render(request, "bulk_upload_books.html", {"form": form})

from django.shortcuts import get_object_or_404, render
from django.template.loader import render_to_string
from django.http import JsonResponse

def book_abstract(request, pk):
    """
    AJAX view: returns JSON {'html': rendered_html} for the book abstract.
    """
    book = get_object_or_404(Book, pk=pk)
    html = render_to_string('book_abstract.html', {'book': book}, request=request)
    return JsonResponse({'html': html})


from django.shortcuts import render
from django.core.mail import send_mail

def report_issue_view(request):
    if request.method == "POST":
        book_id = request.POST.get("book_id")
        student_name = request.POST.get("student_name")
        issue_type = request.POST.get("issue_type")
        student_email = request.POST.get("student_email")

        # Email subject
        subject = f"📚 Book Issue Reported - ID {book_id}"

        # Plain text fallback
        message = f"""
Dear {student_name},

We received your report.

Book ID: {book_id}
Issue: {issue_type}

Our librarian will review it soon.
"""

        # HTML email
        html_message = f"""
<html>
  <body style="font-family: Arial, sans-serif; line-height:1.6; color: #333;">
    <h2 style="color: #2E86C1;">📚 Book Issue Reported</h2>
    <p>Dear <strong>{student_name}</strong>,</p>
    <p>We received your report. Details are as follows:</p>
    <ul>
      <li><strong>Book ID:</strong> {book_id}</li>
      <li><strong>Issue:</strong> {issue_type}</li>
    </ul>
    <p>Our librarian will review it soon and get back to you.</p>
    <p style="color: #555;">Thank you,<br/>Library Team</p>
  </body>
</html>
"""

        from_email = "sarangvss06@gmail.com"
        recipient_list = ["vinothkumars372@gmail.com"]

        # Send email using Twilio SendGrid
        send_mail(
            subject,
            message,
            from_email,
            recipient_list,
            html_message=html_message
        )

        return render(request, "report_issue.html", {"success": True})

    return render(request, "report_issue.html")


from django.shortcuts import render
from .models import Borrow

def search_borrow(request):
    records = None
    error = None
    entered_id = None

    if request.method == "POST":
        entered_id = request.POST.get("student_id", "").strip()

        if entered_id:
            records = Borrow.objects.filter(student_id=entered_id)

            if not records.exists():
                error = f"No records found for Student ID {entered_id}"
        else:
            error = "Please enter a Student ID."

    return render(request, "search_borrow.html", {"records": records, "error": error, "entered_id": entered_id})
